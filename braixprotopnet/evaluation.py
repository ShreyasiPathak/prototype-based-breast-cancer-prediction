import torch
import numpy as np
import openpyxl as op
import matplotlib.pyplot as plt
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from sklearn import metrics

def results_store_excel(train_res, val_res, test_res, per_model_metrics, info_train, info_val, epoch, lr, num_nonzero_prototypes, path_to_results):
    if train_res:
        lines_train = [epoch, lr]
        total_images_train = info_train['total_image']
        avg_train_loss = info_train['avg_loss']
        lines_train.extend([avg_train_loss, info_train['loss_crossent_proto'], info_train['loss_cluster'], info_train['loss_separation'], info_train['loss_crossent_global'], info_train['kd_loss'], info_train['loss_avg_separation']])

        #protopnet branch performance
        correct_train_proto = info_train['correct_proto']
        conf_mat_train_proto = info_train['conf_mat_proto']
        accuracy_train_proto = correct_train_proto / total_images_train
        speci_train_proto = conf_mat_train_proto[0,0]/sum(conf_mat_train_proto[0,:])
        recall_train_proto = conf_mat_train_proto[1,1]/sum(conf_mat_train_proto[1,:])
        prec_train_proto = conf_mat_train_proto[1,1]/sum(conf_mat_train_proto[:,1])
        f1_train_proto = 2*recall_train_proto*prec_train_proto/(recall_train_proto + prec_train_proto)
        prec_train_neg_proto = conf_mat_train_proto[0,0]/sum(conf_mat_train_proto[:,0])
        recall_train_neg_proto = conf_mat_train_proto[0,0]/sum(conf_mat_train_proto[0,:])
        f1_train_neg_proto = 2*recall_train_neg_proto*prec_train_neg_proto/(recall_train_neg_proto + prec_train_neg_proto)
        f1macro_train_proto = (f1_train_proto + f1_train_neg_proto)/2
        lines_train.extend([accuracy_train_proto, f1macro_train_proto, recall_train_proto, speci_train_proto, info_train['auc_proto']])

        #globalnet branch performance
        correct_train_global = info_train['correct_global']
        conf_mat_train_global = info_train['conf_mat_global']
        accuracy_train_global = correct_train_global / total_images_train
        speci_train_global = conf_mat_train_global[0,0]/sum(conf_mat_train_global[0,:])
        recall_train_global = conf_mat_train_global[1,1]/sum(conf_mat_train_global[1,:])
        prec_train_global = conf_mat_train_global[1,1]/sum(conf_mat_train_global[:,1])
        f1_train_global = 2*recall_train_global*prec_train_global/(recall_train_global + prec_train_global)
        prec_train_neg_global = conf_mat_train_global[0,0]/sum(conf_mat_train_global[:,0])
        recall_train_neg_global = conf_mat_train_global[0,0]/sum(conf_mat_train_global[0,:])
        f1_train_neg_global = 2*recall_train_neg_global*prec_train_neg_global/(recall_train_neg_global + prec_train_neg_global)
        f1macro_train_global = (f1_train_global + f1_train_neg_global)/2
        lines_train.extend([accuracy_train_global, f1macro_train_global, recall_train_global, speci_train_global, info_train['auc_global']])

        #both branch performance
        correct_train_both = info_train['correct_both']
        conf_mat_train_both = info_train['conf_mat_both']
        accuracy_train_both = correct_train_both / total_images_train
        speci_train_both = conf_mat_train_both[0,0]/sum(conf_mat_train_both[0,:])
        recall_train_both = conf_mat_train_both[1,1]/sum(conf_mat_train_both[1,:])
        prec_train_both = conf_mat_train_both[1,1]/sum(conf_mat_train_both[:,1])
        f1_train_both = 2*recall_train_both * prec_train_both/(recall_train_both + prec_train_both)
        prec_train_neg_both = conf_mat_train_both[0,0]/sum(conf_mat_train_both[:,0])
        recall_train_neg_both = conf_mat_train_both[0,0]/sum(conf_mat_train_both[0,:])
        f1_train_neg_both = 2 * recall_train_neg_both * prec_train_neg_both/(recall_train_neg_both + prec_train_neg_both)
        f1macro_train_both = (f1_train_both + f1_train_neg_both)/2
        lines_train.extend([accuracy_train_both, f1macro_train_both, recall_train_both, speci_train_both, info_train['auc_both']])
        write_results_xlsx(lines_train, path_to_results, 'train_results')

    if val_res:
        lines_val = [epoch, lr]
        total_images_val = info_val['total_image']
        avg_val_loss = info_val['avg_loss']
        lines_val.extend([avg_val_loss, info_val['loss_crossent_proto'], info_val['loss_cluster'], info_val['loss_separation'], info_val['loss_crossent_global'], info_val['kd_loss'], info_val['loss_avg_separation']])

        #protopnet branch performance
        correct_val_proto = info_val['correct_proto']
        conf_mat_val_proto = info_val['conf_mat_proto']
        speci_val_proto = conf_mat_val_proto[0,0]/sum(conf_mat_val_proto[0,:])
        recall_val_proto = conf_mat_val_proto[1,1]/sum(conf_mat_val_proto[1,:])
        prec_val_proto = conf_mat_val_proto[1,1]/sum(conf_mat_val_proto[:,1])
        f1_val_proto = 2*recall_val_proto * prec_val_proto/(recall_val_proto + prec_val_proto)
        accuracy_val_proto = correct_val_proto / total_images_val
        recall_val_neg_proto = conf_mat_val_proto[0,0]/sum(conf_mat_val_proto[0,:])
        prec_val_neg_proto = conf_mat_val_proto[0,0]/sum(conf_mat_val_proto[:,0])
        f1_val_neg_proto = 2*recall_val_neg_proto * prec_val_neg_proto/(recall_val_neg_proto + prec_val_neg_proto)
        f1macro_val_proto = (f1_val_proto + f1_val_neg_proto)/2
        lines_val.extend([accuracy_val_proto, f1macro_val_proto, recall_val_proto, speci_val_proto, info_val['auc_proto'], info_val['auc_proto_wtmacro']])
        
        #globalnet branch performance
        correct_val_global = info_val['correct_global']
        conf_mat_val_global = info_val['conf_mat_global']
        accuracy_val_global = correct_val_global / total_images_val
        speci_val_global = conf_mat_val_global[0,0]/sum(conf_mat_val_global[0,:])
        recall_val_global = conf_mat_val_global[1,1]/sum(conf_mat_val_global[1,:])
        prec_val_global = conf_mat_val_global[1,1]/sum(conf_mat_val_global[:,1])
        f1_val_global = 2 * recall_val_global * prec_val_global/(recall_val_global + prec_val_global)
        prec_val_neg_global = conf_mat_val_global[0,0]/sum(conf_mat_val_global[:,0])
        recall_val_neg_global = conf_mat_val_global[0,0]/sum(conf_mat_val_global[0,:])
        f1_val_neg_global = 2 * recall_val_neg_global * prec_val_neg_global/(recall_val_neg_global + prec_val_neg_global)
        f1macro_val_global = (f1_val_global + f1_val_neg_global)/2
        lines_val.extend([accuracy_val_global, f1macro_val_global, recall_val_global, speci_val_global, info_val['auc_global'], info_val['auc_global_wtmacro']])

        #both branch performance
        correct_val_both = info_val['correct_both']
        conf_mat_val_both = info_val['conf_mat_both']
        accuracy_val_both = correct_val_both / total_images_val
        speci_val_both = conf_mat_val_both[0,0]/sum(conf_mat_val_both[0,:])
        recall_val_both = conf_mat_val_both[1,1]/sum(conf_mat_val_both[1,:])
        prec_val_both = conf_mat_val_both[1,1]/sum(conf_mat_val_both[:,1])
        f1_val_both = 2 * recall_val_both * prec_val_both/(recall_val_both + prec_val_both)
        prec_val_neg_both = conf_mat_val_both[0,0]/sum(conf_mat_val_both[:,0])
        recall_val_neg_both = conf_mat_val_both[0,0]/sum(conf_mat_val_both[0,:])
        f1_val_neg_both = 2 * recall_val_neg_both * prec_val_neg_both/(recall_val_neg_both + prec_val_neg_both)
        f1macro_val_both = (f1_val_both + f1_val_neg_both)/2
        lines_val.extend([accuracy_val_both, f1macro_val_both, recall_val_both, speci_val_both, info_val['auc_both'], info_val['auc_both_wtmacro']])
        
        write_results_xlsx(lines_val, path_to_results, 'val_results')

    if test_res:
        lines_test = [epoch]
        lines_test.extend(per_model_metrics)
    
        write_results_xlsx(lines_test, path_to_results, 'test_results')

def conf_mat_create(predicted, true, correct, total_images, conf_mat, classes):
    total_images+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat_batch=confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    conf_mat=conf_mat+conf_mat_batch
    return correct, total_images, conf_mat, conf_mat_batch

def write_results_xlsx(results, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    sheet = wb[sheetname]
    sheet.append(results)
    wb.save(path_to_results)

def write_results_xlsx_confmat(numclasses, results, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    
    sheet.append(numclasses)
    for row in results.tolist():
        sheet.append(row)
    wb.save(path_to_results)

def aggregate_performance_metrics(numclasses, y_true, y_pred, y_prob): 
    try:
        prec_bin = metrics.precision_score(y_true, y_pred, average = 'binary')
    except:
        prec_bin = 0.0
    precmicro = metrics.precision_score(y_true, y_pred, average = 'micro')
    precmacro = metrics.precision_score(y_true, y_pred, average = 'macro')
    try:
        recall_bin = metrics.recall_score(y_true, y_pred, average = 'binary')
    except:
        recall_bin = 0.0
    recallmicro = metrics.recall_score(y_true, y_pred, average = 'micro')
    recallmacro = metrics.recall_score(y_true, y_pred, average = 'macro')
    try:
        f1_bin = metrics.f1_score(y_true, y_pred, average = 'binary')
    except:
        f1_bin = 0.0
    f1micro = metrics.f1_score(y_true, y_pred, average = 'micro')
    f1macro = metrics.f1_score(y_true, y_pred, average='macro')
    f1wtmacro=metrics.f1_score(y_true, y_pred, average='weighted')
    acc = metrics.accuracy_score(y_true, y_pred)
    cohen_kappa=metrics.cohen_kappa_score(y_true, y_pred)
    try:
        if numclasses > 2:
            auc = metrics.roc_auc_score(y_true, y_prob, average='macro', multi_class='ovo')
            auc_wtmacro = metrics.roc_auc_score(y_true, y_prob, average='weighted', multi_class='ovo')
        else:
            auc = metrics.roc_auc_score(y_true, y_prob)
            auc_wtmacro=0.0
    except:
        auc=0.0
        auc_wtmacro=0.0
    
    each_model_metrics=[prec_bin, precmicro, precmacro, recall_bin, recallmicro, recallmacro, f1_bin, f1micro, f1macro, f1wtmacro, acc, cohen_kappa, auc, auc_wtmacro]
    return each_model_metrics

def classspecific_performance_metrics(numclasses, y_true, y_pred):
    score_dict = classification_report(y_true, y_pred, labels=numclasses, output_dict = True)
    print(score_dict)
    results_all = []
    flag=0
    for key in score_dict.keys():
        if isinstance(score_dict[key], dict):
            if flag == 0:
                results_all.append(['class'] + list(score_dict[key].keys()))
                flag = 1
            results_all.append([key] + list(score_dict[key].values())) 
        else:
            results_all.append([key, score_dict[key]])
    
    print(results_all)
    return results_all

def write_results_classspecific(results_all, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    for result in results_all:
        sheet.append(result)
    wb.save(path_to_results)
