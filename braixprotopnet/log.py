import os
import openpyxl as op
from openpyxl import Workbook

def create_logger(log_filename, display=True):
    f = open(log_filename, 'a')
    counter = [0]
    # this function will still have access to f after create_logger terminates
    def logger(text):
        if display:
            print(text)
        f.write(text + '\n')
        counter[0] += 1
        if counter[0] % 10 == 0:
            f.flush()
            os.fsync(f.fileno())
        # Question: do we need to flush()
    return logger, f.close

def create_results(log_dir, results_file):
    """
    Create a csv for logging information
    :param log_name: The name of the log. The log filename will be <log_name>.csv.
    :param key_name: The name of the attribute that is used as key (e.g. epoch number)
    :param value_names: The names of the attributes that are logged
    """
    if not os.path.exists(log_dir + f'/{results_file}.xlsx'):
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title="train_results"
        header_train = ['Epoch','lr','AvgLoss','CrossEntProtoLoss','ClusterProtoLoss','SepProtoLoss','CrossEntGlobalLoss','KDLoss','AvgSeparationLoss','Accuracy_Proto','F1macro_Proto','Recall_Proto','Speci_Proto','AUC_Proto', 'Accuracy_Global','F1macro_Global','Recall_Global','Speci_Global','AUC_Global','Accuracy_Both','F1macro_Both','Recall_Both','Speci_Both','AUC_Both']
        sheet1.append(header_train)
        sheet2 = wb.create_sheet('val_results')
        header_val = ['Epoch','lr','AvgLoss','CrossEntProtoLoss','ClusterProtoLoss','SepProtoLoss','CrossEntGlobalLoss','KDLoss','AvgSeparationLoss','Accuracy_Proto','F1macro_Proto','Recall_Proto','Speci_Proto','AUC_Proto','AUC_WtMacro_Proto', 'Accuracy_Global','F1macro_Global','Recall_Global','Speci_Global','AUC_Global','AUC_WtMacro_Global','Accuracy_Both','F1macro_Both','Recall_Both','Speci_Both','AUC_Both','AUC_WtMacro_Both']
        sheet2.append(header_val)
        sheet3 = wb.create_sheet('confmat_train_val_test')
        sheet3 = wb.create_sheet('test_results') 
        header1 = ['Epoch','AvgLoss','CrossEntProtoLoss','ClusterProtoLoss','SepProtoLoss','CrossEntGlobalLoss','KDLoss','AvgSeparationLoss','L1']
        header1 = header1 + ['PrecisionBinProto','PrecisionMicroProto','PrecisionMacroProto','RecallBinProto','RecallMicroProto','RecallMacroProto','F1BinProto','F1MicroProto','F1macroProto','F1wtmacroProto','AccProto','CohensKappaProto','AUCProto','AUCWtMacroProto']
        header1 = header1 + ['PrecisionBinGlobal','PrecisionMicroGlobal','PrecisionMacroGlobal','RecallBinGlobal','RecallMicroGlobal','RecallMacroGlobal','F1BinGlobal','F1MicroGlobal','F1macroGlobal','F1wtmacroGlobal','AccGlobal','CohensKappaGlobal','AUCGlobal','AUCWtMacroGlobal']
        header1 = header1 + ['PrecisionBinBoth','PrecisionMicroBoth','PrecisionMacroBoth','RecallBinBoth','RecallMicroBoth','RecallMacroBoth','F1BinBoth','F1MicroBoth','F1macroBoth','F1wtmacroBoth','AccBoth','CohensKappaBoth','AUCBoth','AUCWtMacroBoth']
        sheet3.append(header1)
        sheet4 = wb.create_sheet('classpecific_test_results') 
        wb.save(log_dir + f'/{results_file}.xlsx')

