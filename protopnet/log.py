import os
import openpyxl as op
from openpyxl import Workbook

def create_logger(log_filename, display=True):
    f = open(log_filename, 'a')
    counter = [0]
    # this function will still have access to f after create_logger terminates
    def logger(text):
        if display:
            print(text)
        f.write(text + '\n')
        counter[0] += 1
        if counter[0] % 10 == 0:
            f.flush()
            os.fsync(f.fileno())
        # Question: do we need to flush()
    return logger, f.close

def create_results(log_dir, results_file):
    """
    Create a csv for logging information
    :param log_name: The name of the log. The log filename will be <log_name>.csv.
    :param key_name: The name of the attribute that is used as key (e.g. epoch number)
    :param value_names: The names of the attributes that are logged
    """
    if not os.path.exists(log_dir + f'/{results_file}.xlsx'):
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title="train_results"
        header=['Epoch','lr','AvgLoss','CrossEntLoss','ClusterLoss','SepLoss','AvgSeparationLoss','L1','Accuracy','F1macro','Recall','Speci','AUC']
        sheet1.append(header)
        sheet2 = wb.create_sheet('val_results')
        sheet2.append(header)
        sheet2 = wb.create_sheet('confmat_train_val_test')
        sheet3 = wb.create_sheet('test_results') 
        header1 = ['Epoch','AvgLoss','CrossEntLoss','ClusterLoss','SepLoss','AvgSeparationLoss','L1']
        header1 = header1 + ['PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','CohensKappa','AUC','AUCWtMacro']
        sheet3.append(header1)
        wb.save(log_dir + f'/{results_file}.xlsx')

