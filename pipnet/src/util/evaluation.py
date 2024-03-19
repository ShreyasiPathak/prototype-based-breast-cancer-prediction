import torch
import numpy as np
import openpyxl as op
import matplotlib.pyplot as plt
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from sklearn import metrics

def results_store_excel(train_res, val_res, test_res, per_model_metrics, info_train, info_val, epoch, lr, num_nonzero_prototypes, path_to_results):
    #train_res, val_res, test_res, per_model_metrics, correct_train, total_images_train, avg_train_loss, correct_test, total_images_test, avg_test_loss_tanh, avg_test_loss_class, epoch, conf_mat_train, conf_mat_test, lr, auc_val, num_nonzero_prototypes, path_to_results
    lines = [epoch, lr]
    if train_res:
        total_images_train = info_train['total_images_train']
        avg_train_loss = info_train['loss']
        lines.extend([avg_train_loss, info_train['align_loss'], info_train['tanh_loss'], info_train['class_loss']])

        #pipnet metrics
        correct_train = info_train['correct_train']
        conf_mat_train = info_train['conf_mat_train']
        accuracy_train = correct_train / total_images_train
        speci_train = conf_mat_train[0,0]/sum(conf_mat_train[0,:])
        recall_train = conf_mat_train[1,1]/sum(conf_mat_train[1,:])
        prec_train = conf_mat_train[1,1]/sum(conf_mat_train[:,1])
        f1_train = 2*recall_train*prec_train/(recall_train+prec_train)
        prec_train_neg = conf_mat_train[0,0]/sum(conf_mat_train[:,0])
        recall_train_neg = conf_mat_train[0,0]/sum(conf_mat_train[0,:])
        f1_train_neg = 2*recall_train_neg*prec_train_neg/(recall_train_neg+prec_train_neg)
        f1macro_train = (f1_train+f1_train_neg)/2
        lines.extend([accuracy_train, f1macro_train, recall_train, speci_train])

    if val_res:
        total_images_val = info_val['total_images_test']
        avg_val_loss = info_val['loss']
        lines.extend([avg_val_loss, info_val['tanh_loss'], info_val['class_loss']])

        correct_val = info_val['correct_test']
        conf_mat_val = info_val['confusion_matrix']
        speci_val = conf_mat_val[0,0]/sum(conf_mat_val[0,:])
        recall_val = conf_mat_val[1,1]/sum(conf_mat_val[1,:])
        prec_val = conf_mat_val[1,1]/sum(conf_mat_val[:,1])
        f1_val = 2*recall_val*prec_val/(recall_val+prec_val)
        accuracy_val = correct_val / total_images_val
        recall_val_neg = conf_mat_val[0,0]/sum(conf_mat_val[0,:])
        prec_val_neg = conf_mat_val[0,0]/sum(conf_mat_val[:,0])
        f1_val_neg = 2*recall_val_neg*prec_val_neg/(recall_val_neg+prec_val_neg)
        f1macro_val = (f1_val + f1_val_neg)/2
        lines.extend([accuracy_val, f1macro_val, recall_val, speci_val, info_val['AUC_test'], info_val['AUC_weightedmacro_test'], num_nonzero_prototypes])

    if test_res:
        lines.extend(per_model_metrics)
    
    write_results_xlsx(lines, path_to_results, 'train_val_results')

def conf_mat_create(predicted, true, correct, total_images, conf_mat, classes):
    total_images+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat_batch=confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    conf_mat=conf_mat+conf_mat_batch
    return correct, total_images, conf_mat, conf_mat_batch

def write_results_xlsx(results, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    sheet = wb[sheetname]
    sheet.append(results)
    wb.save(path_to_results)

def write_results_xlsx_confmat(numclasses, results, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    
    sheet.append(numclasses)
    for row in results.tolist():
        sheet.append(row)
    wb.save(path_to_results)

def aggregate_performance_metrics(numclasses, y_true, y_pred, y_prob): 
    if numclasses > 2:
        prec_bin = 0.0
    else:
        prec_bin = metrics.precision_score(y_true, y_pred, average = 'binary')
    precmicro = metrics.precision_score(y_true, y_pred, average = 'micro')
    precmacro = metrics.precision_score(y_true, y_pred, average = 'macro')
    if numclasses > 2:
        recall_bin = 0.0
    else:
        recall_bin = metrics.recall_score(y_true, y_pred, average = 'binary')
    recallmicro = metrics.recall_score(y_true, y_pred, average = 'micro')
    recallmacro = metrics.recall_score(y_true, y_pred, average = 'macro')
    if numclasses > 2:
        f1_bin = 0.0
    else:
        f1_bin = metrics.f1_score(y_true, y_pred, average = 'binary')
    f1micro = metrics.f1_score(y_true, y_pred, average = 'micro')
    f1macro = metrics.f1_score(y_true, y_pred, average='macro')
    f1wtmacro=metrics.f1_score(y_true, y_pred, average='weighted')
    acc = metrics.accuracy_score(y_true, y_pred)
    cohen_kappa=metrics.cohen_kappa_score(y_true, y_pred)
    try:
        if numclasses > 2:
            auc=metrics.roc_auc_score(y_true, y_prob, multi_class='ovo') #changed ovr to ovo
            auc_wtmacro = metrics.roc_auc_score(y_true, y_prob, average='weighted', multi_class='ovo')
        else:
            auc=metrics.roc_auc_score(y_true, y_prob)
            auc_wtmacro=0.0
    except:
        auc=0.0
        auc_wtmacro=0.0
    
    each_model_metrics=[prec_bin, precmicro, precmacro, recall_bin, recallmicro, recallmacro, f1_bin, f1micro, f1macro, f1wtmacro, acc, cohen_kappa, auc, auc_wtmacro]
    return each_model_metrics

def classspecific_performance_metrics(numclasses, y_true, y_pred):
    score_dict = classification_report(y_true, y_pred, labels=numclasses, output_dict = True)
    print(score_dict)
    results_all = []
    flag=0
    for key in score_dict.keys():
        if isinstance(score_dict[key], dict):
            if flag == 0:
                results_all.append(['class'] + list(score_dict[key].keys()))
                flag = 1
            results_all.append([key] + list(score_dict[key].values())) 
        else:
            results_all.append([key, score_dict[key]])
    
    print(results_all)
    return results_all

def write_results_classspecific(results_all, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    for result in results_all:
        sheet.append(result)
