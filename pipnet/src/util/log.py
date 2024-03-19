import os
import argparse
import openpyxl as op
from openpyxl import Workbook

from util.args import save_args, load_args

from util import evaluation

class Log:

    """
    Object for managing the log directory
    """

    def __init__(self, log_dir: str):  # Store log in log_dir

        self._log_dir = log_dir
        self._logs = dict()

        # Ensure the directories exist
        if not os.path.isdir(self.log_dir):
            os.mkdir(self.log_dir)
        if not os.path.isdir(self.metadata_dir):
            os.mkdir(self.metadata_dir)
        if not os.path.isdir(self.checkpoint_dir):
            os.mkdir(self.checkpoint_dir)
        

    @property
    def log_dir(self):
        return self._log_dir

    @property
    def checkpoint_dir(self):
        return self._log_dir + '/checkpoints'

    @property
    def metadata_dir(self):
        return self._log_dir + '/metadata'

    def log_message(self, msg: str):
        """
        Write a message to the log file
        :param msg: the message string to be written to the log file
        """
        if not os.path.isfile(self.log_dir + '/log.txt'):
            open(self.log_dir + '/log.txt', 'w').close() #make log file empty if it already exists
        with open(self.log_dir + '/log.txt', 'a') as f:
            f.write(msg+"\n")

    def create_log(self, log_name: str, results_file: str, key_name: str, *value_names):
        """
        Create a csv for logging information
        :param log_name: The name of the log. The log filename will be <log_name>.csv.
        :param key_name: The name of the attribute that is used as key (e.g. epoch number)
        :param value_names: The names of the attributes that are logged
        """
        if log_name in self._logs.keys():
            raise Exception('Log already exists!')
        # Add to existing logs
        self._logs[log_name] = (key_name, value_names)
        # Create log file. Create columns
        with open(self.log_dir + f'/{log_name}.csv', 'a') as f:
            f.write(','.join((key_name,) + value_names) + '\n')
        if not os.path.exists(self.log_dir + f'/{results_file}.xlsx'):
            #with open(self.log_dir + f'/{results_file}.xlsx', 'a') as f:
            wb = Workbook()
            sheet1 = wb.active
            sheet1.title="train_val_results"
            header=['Epoch', 'lr', 'AvgLossTrain', 'AlignLossTrain', 'TanhLossTrain', 'CELossTrain', 'AccuracyTrain', 'F1macroTrain', 'RecallTrain', 'SpeciTrain', 'AvgLossVal', 'TanhLossVal', 'CELossVal', 'AccuracyVal', 'F1macroVal', 'RecallVal', 'SpeciVal', 'AUCVal', 'AUCWtMacroVal','num_nonzero_prototypes']
            sheet1.append(header)
            sheet2 = wb.create_sheet('confmat_train_val_test')
            sheet3 = wb.create_sheet('test_results') 
            sheet3.append(['Epoch', 'AvgLoss', 'TanhLoss', 'CELoss', 'PrecisionBin', 'PrecisionMicro', 'PrecisionMacro', 'RecallBin', 'RecallMicro', 'RecallMacro', 'F1Bin', 'F1Micro', 'F1macro', 'F1wtmacro', 'Acc', 'Cohens Kappa', 'AUC', 'AUCWtMacro', 'num_nonzero_prototypes'])
            sheet4 = wb.create_sheet('metrics_view_wise')
            wb.save(self.log_dir + f'/{results_file}.xlsx')

    def log_values(self, log_name, key, *values):
        """
        Log values in an existent log file
        :param log_name: The name of the log file
        :param key: The key attribute for logging these values
        :param values: value attributes that will be stored in the log
        """
        if log_name not in self._logs.keys():
            raise Exception('Log not existent!')
        if len(values) != len(self._logs[log_name][1]):
            raise Exception('Not all required values are logged!')
        # Write a new line with the given values
        with open(self.log_dir + f'/{log_name}.csv', 'a') as f:
            f.write(','.join(str(v) for v in (key,) + values) + '\n')

    def log_args(self, args: argparse.Namespace):
        save_args(args, self._log_dir)

